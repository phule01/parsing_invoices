from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from database import get_db
from models import Invoice, InvoiceItem, Product
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

router = APIRouter()

@router.get("/summary")
def get_summary(db: Session = Depends(get_db)):
    """Get sales summary report"""
    total_invoices = db.query(func.count(Invoice.id)).scalar()
    total_revenue = db.query(func.sum(Invoice.total_amount)).scalar() or 0
    total_items = db.query(func.sum(InvoiceItem.quantity)).scalar() or 0
    
    return {
        "total_invoices": total_invoices,
        "total_revenue": float(total_revenue),
        "total_items_sold": total_items
    }

@router.get("/by-product")
def get_sales_by_product(db: Session = Depends(get_db)):
    """Get sales breakdown by product"""
    results = db.query(
        Product.name,
        func.sum(InvoiceItem.quantity).label("quantity_sold"),
        func.sum(InvoiceItem.total_price).label("revenue")
    ).join(InvoiceItem).group_by(Product.id, Product.name).all()
    
    return [
        {
            "product_name": r[0],
            "quantity_sold": r[1] or 0,
            "revenue": float(r[2] or 0)
        }
        for r in results
    ]

@router.get("/by-user")
def get_sales_by_user(db: Session = Depends(get_db)):
    """Get sales breakdown by user"""
    results = db.query(
        Invoice.user_id,
        func.count(Invoice.id).label("num_invoices"),
        func.sum(Invoice.total_amount).label("total_spent")
    ).group_by(Invoice.user_id).all()
    
    return [
        {
            "user_id": r[0],
            "num_invoices": r[1],
            "total_spent": float(r[2] or 0)
        }
        for r in results
    ]

@router.get("/export-excel")
def export_invoices_excel(db: Session = Depends(get_db)):
    """Export all invoices to Excel file"""
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"
    
    # Headers  
    headers = ["Invoice #", "Supplier", "Tax Code", "Date", "Total", "VAT", "Status", "Source"]
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Get invoices
    invoices = db.query(Invoice).all()
    for inv in invoices:
        ws.append([
            inv.invoice_number or "",
            inv.seller_name or "",
            inv.seller_tax_code or "",
            str(inv.invoice_date) if inv.invoice_date else "",
            float(inv.total_amount or 0),
            float(inv.vat_amount or 0),
            inv.status or "",
            inv.source_type or ""
        ])
    
    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if len(str(cell.value or "")) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column_letter].width = max_length + 2
    
    # Create items sheet
    ws_items = wb.create_sheet("Items")
    ws_items.append(["Invoice #", "Product", "Unit", "Quantity", "Unit Price", "Total", "VAT %"])
    
    items = db.query(InvoiceItem, Invoice.invoice_number).join(
        Invoice, InvoiceItem.invoice_id == Invoice.id
    ).all()
    
    for item, inv_num in items:
        ws_items.append([
            inv_num,
            item.item_name or "",
            item.unit or "",
            float(item.quantity or 0),
            float(item.unit_price or 0),
            float(item.total_price or 0),
            float(item.vat_rate or 0)
        ])
    
    # Auto-width items columns
    for column in ws_items.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if len(str(cell.value or "")) > max_length:
                max_length = len(str(cell.value))
        ws_items.column_dimensions[column_letter].width = max_length + 2
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return as streaming response (FileResponse requires a path; BytesIO needs StreamingResponse)
    filename = f"invoices_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )